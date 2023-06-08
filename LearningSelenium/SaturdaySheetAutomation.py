import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# Load the Excel file
workbook = openpyxl.load_workbook("C:\Python-Selenium\PythonSeleniumProject1\LearningSelenium\Excel.xlsx")
sheet = workbook["Saturday"]

# Iterate over each row
for i in range(3, sheet.max_row + 1):
    # Read data from desired columns (assuming column indexes are 0 and 1)
    searchData1 = sheet.cell(row=i, column=3).value

    # Set up ChromeDriver
    # service = Service(ChromeDriverManager().install())
    # driver = webdriver.Chrome(service=service)
    driver = webdriver.Chrome("C:\BrowsersDrivers\chromedriver_win32\chromedriver.exe")
    driver.maximize_window()

    # Perform Google search
    driver.get("https://www.google.com/")
    search_box = driver.find_element(By.NAME, "q")
    search_box.send_keys(searchData1)
    # search_box.submit()
    driver.implicitly_wait(5)

    # Finding the Suggestion Elements
    suggestion_list = driver.find_elements(By.XPATH, "//ul[@role='listbox']//li/descendant::div[@class='wM6W7d']")

    suggestion_texts = []
    for suggestion in suggestion_list:
        suggestion_texts.append(suggestion.text)

    # Print largest and smallest suggestion text
    largest_suggestion = max(suggestion_texts)
    smallest_suggestion = min(suggestion_texts)

    print("Largest Suggestion:", largest_suggestion)
    print("Smallest Suggestion:", smallest_suggestion)

    # Write largest and smallest suggestion to Excel file (assuming columns 4 and 5)
    sheet.cell(row=i, column=4).value = largest_suggestion
    sheet.cell(row=i, column=5).value = smallest_suggestion

    # Clear search box for the next iteration
    search_box.clear()
    driver.quit()

# Save the changes to the Excel file
workbook.save("C:\Python-Selenium\PythonSeleniumProject1\LearningSelenium\Excel.xlsx")
workbook.close()