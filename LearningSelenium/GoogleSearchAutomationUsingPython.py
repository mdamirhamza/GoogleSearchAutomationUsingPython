import openpyxl #import openpyxl for read and write to excel file
from selenium import webdriver
from selenium.webdriver.common.by import By

# Load the Excel file
workbook = openpyxl.load_workbook("C:\Python-Selenium\PythonSeleniumProject1\LearningSelenium\Excel.xlsx")

# Array of desired sheet names
desired_sheet_names = ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

# Iterate through each desired sheet name
for sheet_name in desired_sheet_names:

    # Check if the sheet exists
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        row_count = sheet.max_row - 1  # skip header row

        # Iterate over each row
        for i in range(3, row_count + 1):

            # Read data from desired columns
            search_data1 = sheet.cell(row=i, column=3).value

            # Set up ChromeDriver
            driver = webdriver.Chrome("C:\BrowsersDrivers\chromedriver_win32\chromedriver.exe")
            driver.maximize_window()

            # Perform Google search
            driver.get("https://www.google.com/")
            search_box = driver.find_element(By.NAME, "q")
            search_box.send_keys(search_data1)
            driver.implicitly_wait(2)

            # Finding the Suggestion Elements
            suggestion_list = driver.find_elements(By.XPATH, "//ul[@role='listbox']//li/descendant::div[@class='wM6W7d']")

            suggestion_texts = []
            for suggestion in suggestion_list:
                suggestion_texts.append(suggestion.text)

            # Print largest and smallest suggestion text
            largest_suggestion = max(suggestion_texts)
            smallest_suggestion = min(suggestion_texts)

            print("Largest Suggestion : ", largest_suggestion)
            print("Smallest Suggestion : ", smallest_suggestion)

            # Write largest and smallest suggestion to Excel file (assuming columns 4 and 5)
            sheet.cell(row=i, column=4).value = largest_suggestion
            sheet.cell(row=i, column=5).value = smallest_suggestion

            search_box.clear()
            driver.quit()

# Save to the Excel file
workbook.save("C:\Python-Selenium\PythonSeleniumProject1\LearningSelenium\Excel.xlsx")
workbook.close()